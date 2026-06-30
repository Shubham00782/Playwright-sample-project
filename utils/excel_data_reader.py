import openpyxl


class ExcelUtils:
    """
    A utility class for interacting with Excel files using openpyxl.
    This class provides static methods for Data Driven Testing (DDT) operations such as:
    1) Counting the maximum number of rows
    2) Counting the maximum number of columns
    3) Reading data from a specific cell
    4) Writing data to a specific cell
    """

    @staticmethod
    def row_count(files, sheet):
        """
        Count the maximum number of rows in the given Excel sheet.

        :param files: The file path to the Excel workbook.
        :param sheet: The name of the sheet to count rows in.
        :return: The maximum number of rows in the sheet.
        """
        workbook = openpyxl.load_workbook(files)
        sheet = workbook[sheet]
        rows = sheet.max_row
        return rows

    @staticmethod
    def col_count(files, sheet):
        """
        Count the maximum number of columns in the given Excel sheet.

        :param files: The file path to the Excel workbook.
        :param sheet: The name of the sheet to count columns in.
        :return: The maximum number of columns in the sheet.
        """
        workbook = openpyxl.load_workbook(files)
        sheet = workbook[sheet]
        cols = sheet.max_column
        return cols

    @staticmethod
    def read_data(files, sheet, cell_chords):
        """
        Read data from a specific cell in the given Excel sheet.

        :param files: The file path to the Excel workbook.
        :param sheet: The name of the sheet to read data from.
        :param cell_chords: The cell coordinates (e.g., 'A1') to read data from.
        :return: The value contained in the specified cell.
        """
        workbook = openpyxl.load_workbook(files)
        sheet = workbook[sheet]
        return sheet[cell_chords].value

    @staticmethod
    def write_data(files, sheet, cell_chords, data):
        """
        Write data to a specific cell in the given Excel sheet.

        :param files: The file path to the Excel workbook.
        :param sheet: The name of the sheet to write data to.
        :param cell_chords: The cell coordinates (e.g., 'A1') to write data to.
        :param data: The data to write into the specified cell.
        """
        workbook = openpyxl.load_workbook(files)
        sheet = workbook[sheet]
        sheet[cell_chords].value = data
        workbook.save(files)

    @staticmethod
    def get_excel_data(files, sheet_name):
        workbook = openpyxl.load_workbook(files)
        sheet = workbook[sheet_name]

        data = []

        headers = [
            sheet.cell(row=1, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        for row in range(2, sheet.max_row + 1):
            row_data = {}

            for col in range(1, sheet.max_column + 1):
                row_data[headers[col - 1]] = sheet.cell(
                    row=row,
                    column=col
                ).value

            data.append(row_data)

        return data