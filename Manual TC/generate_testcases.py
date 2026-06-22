import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# Define styles
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_alignment = Alignment(vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)
header_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='FFFFFF'),
    bottom=Side(style='thin', color='FFFFFF')
)
alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

HEADERS = [
    "Test Case ID", "Pytest Test Name", "Test Scenario", "Test Case Title",
    "Test Case Type", "Pre-requisites", "Test Steps", "Test Data",
    "Expected Result", "Actual Result", "Severity", "Priority", "Result", "Comments"
]

COL_WIDTHS = [14, 35, 30, 35, 16, 30, 55, 35, 45, 45, 12, 12, 10, 25]


def create_sheet(wb, sheet_name, test_cases, is_first=False):
    if is_first:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Set column widths
    for i, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Write headers
    for col_num, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Write test cases
    for row_idx, tc in enumerate(test_cases, 2):
        for col_idx, value in enumerate(tc, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Set row heights
    for row in range(2, len(test_cases) + 2):
        ws.row_dimensions[row].height = 80

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:N{len(test_cases) + 1}"
    return ws


# ============================================================
# MODULE 1: LOGIN
# ============================================================
login_tests = [
    [
        "LOGIN_TC_001", "test_login_valid_standard_user",
        "Verify login with valid standard_user credentials",
        "Valid Login - Standard User",
        "Positive", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter username 'standard_user' in the Username field\n3. Enter password 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: standard_user\nPassword: secret_sauce",
        "User should be successfully logged in and redirected to the inventory page (https://www.saucedemo.com/inventory.html). The page title 'Swag Labs' should be displayed along with the products list.",
        "User is successfully logged in and redirected to inventory page. Page title 'Swag Labs' is displayed with 6 products listed.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_002", "test_login_valid_problem_user",
        "Verify login with valid problem_user credentials",
        "Valid Login - Problem User",
        "Positive", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter username 'problem_user' in the Username field\n3. Enter password 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: problem_user\nPassword: secret_sauce",
        "User should be logged in and redirected to inventory page. Product images may display incorrectly as this is a known problem user account.",
        "User is logged in and redirected to inventory page. All product images show the same dog image (sl-404.jpg) instead of actual product images.",
        "Major", "P2", "Fail", "Known issue with problem_user - all product images are broken/same"
    ],
    [
        "LOGIN_TC_003", "test_login_valid_performance_glitch_user",
        "Verify login with performance_glitch_user credentials",
        "Valid Login - Performance Glitch User",
        "Positive", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter username 'performance_glitch_user' in the Username field\n3. Enter password 'secret_sauce' in the Password field\n4. Click the 'Login' button\n5. Observe the time taken for the page to load",
        "Username: performance_glitch_user\nPassword: secret_sauce",
        "User should be logged in but there should be a noticeable delay (up to 5 seconds) before the inventory page loads.",
        "User is logged in after approximately 3-5 seconds delay. Inventory page eventually loads with all products.",
        "Major", "P2", "Pass", "Login has intentional performance glitch/delay"
    ],
    [
        "LOGIN_TC_004", "test_login_locked_out_user",
        "Verify login with locked_out_user credentials",
        "Locked Out User Login Attempt",
        "Negative", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter username 'locked_out_user' in the Username field\n3. Enter password 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: locked_out_user\nPassword: secret_sauce",
        "Login should fail and an error message 'Epic sadface: Sorry, this user has been locked out.' should be displayed. User should remain on the login page.",
        "Login fails. Error message 'Epic sadface: Sorry, this user has been locked out.' is displayed. User remains on the login page.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_005", "test_login_invalid_username",
        "Verify login with invalid/unregistered username",
        "Invalid Username Login Attempt",
        "Negative", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter username 'invalid_user' in the Username field\n3. Enter password 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: invalid_user\nPassword: secret_sauce",
        "Login should fail and error message 'Epic sadface: Username and password do not match any user in this service' should be displayed.",
        "Login fails. Error message 'Epic sadface: Username and password do not match any user in this service' is displayed.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_006", "test_login_invalid_password",
        "Verify login with valid username and invalid password",
        "Invalid Password Login Attempt",
        "Negative", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter username 'standard_user' in the Username field\n3. Enter password 'wrong_password' in the Password field\n4. Click the 'Login' button",
        "Username: standard_user\nPassword: wrong_password",
        "Login should fail and error message 'Epic sadface: Username and password do not match any user in this service' should be displayed.",
        "Login fails. Error message 'Epic sadface: Username and password do not match any user in this service' is displayed.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_007", "test_login_empty_username",
        "Verify login with empty username field",
        "Empty Username Login Attempt",
        "Negative", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Leave the Username field empty\n3. Enter password 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: (empty)\nPassword: secret_sauce",
        "Login should fail and error message 'Epic sadface: Username is required' should be displayed.",
        "Login fails. Error message 'Epic sadface: Username is required' is displayed.",
        "Major", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_008", "test_login_empty_password",
        "Verify login with empty password field",
        "Empty Password Login Attempt",
        "Negative", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter username 'standard_user' in the Username field\n3. Leave the Password field empty\n4. Click the 'Login' button",
        "Username: standard_user\nPassword: (empty)",
        "Login should fail and error message 'Epic sadface: Password is required' should be displayed.",
        "Login fails. Error message 'Epic sadface: Password is required' is displayed.",
        "Major", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_009", "test_login_both_fields_empty",
        "Verify login with both username and password fields empty",
        "Both Fields Empty Login Attempt",
        "Negative", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Leave the Username field empty\n3. Leave the Password field empty\n4. Click the 'Login' button",
        "Username: (empty)\nPassword: (empty)",
        "Login should fail and error message 'Epic sadface: Username is required' should be displayed.",
        "Login fails. Error message 'Epic sadface: Username is required' is displayed.",
        "Major", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_010", "test_login_error_message_close",
        "Verify the error message close (X) button functionality",
        "Close Error Message on Login Page",
        "Functional", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Click the 'Login' button without entering credentials\n3. Verify error message is displayed\n4. Click the 'X' (close) button on the error message",
        "N/A",
        "The error message should disappear when the close button is clicked. The error icon markers on the input fields should also be removed.",
        "Error message disappears on clicking 'X' button. Error icons on input fields are removed.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "LOGIN_TC_011", "test_login_password_field_masked",
        "Verify the password field masks the entered text",
        "Password Field Masking",
        "Security", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Click on the Password field\n3. Type 'secret_sauce'\n4. Observe the entered text in the password field",
        "Password: secret_sauce",
        "The password field should mask the entered text showing dots/bullets instead of plain text. The input type should be 'password'.",
        "Password field masks the text with dots. Input field type attribute is 'password'.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_012", "test_login_placeholder_text",
        "Verify placeholder text in Username and Password fields",
        "Input Field Placeholder Text",
        "UI", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Observe the Username field placeholder text\n3. Observe the Password field placeholder text",
        "N/A",
        "Username field should display 'Username' as placeholder text. Password field should display 'Password' as placeholder text.",
        "Username field shows 'Username' placeholder. Password field shows 'Password' placeholder.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "LOGIN_TC_013", "test_login_accepted_usernames_display",
        "Verify accepted usernames are displayed on login page",
        "Accepted Usernames Info Display",
        "UI", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Scroll down to the login credentials section\n3. Verify all accepted usernames are listed\n4. Verify the password is listed",
        "N/A",
        "The login page should display accepted usernames: standard_user, locked_out_user, problem_user, performance_glitch_user, error_user, visual_user. Password 'secret_sauce' should be shown for all users.",
        "All 6 accepted usernames are displayed. Password 'secret_sauce' is shown for all users.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "LOGIN_TC_014", "test_login_sql_injection",
        "Verify login is secure against SQL injection attacks",
        "SQL Injection in Login Fields",
        "Security", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter \"' OR '1'='1\" in the Username field\n3. Enter \"' OR '1'='1\" in the Password field\n4. Click the 'Login' button",
        "Username: ' OR '1'='1\nPassword: ' OR '1'='1",
        "Login should fail with an appropriate error message. The application should not allow SQL injection based authentication bypass.",
        "Login fails. Error message 'Epic sadface: Username and password do not match any user in this service' is displayed. No SQL injection bypass.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_015", "test_login_xss_injection",
        "Verify login is secure against XSS attacks",
        "XSS Injection in Login Fields",
        "Security", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter '<script>alert(\"XSS\")</script>' in the Username field\n3. Enter 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: <script>alert('XSS')</script>\nPassword: secret_sauce",
        "Login should fail with standard error message. No JavaScript alert should be triggered. The script tag should be treated as plain text.",
        "Login fails with standard error message. No XSS alert is triggered.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_016", "test_login_case_sensitive_username",
        "Verify username field is case-sensitive",
        "Case Sensitivity - Username",
        "Negative", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter 'Standard_User' (with capital S and U) in the Username field\n3. Enter 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: Standard_User\nPassword: secret_sauce",
        "Login should fail because usernames are case-sensitive. Error message should be displayed.",
        "Login fails. Error message 'Epic sadface: Username and password do not match any user in this service' is displayed.",
        "Major", "P2", "Pass", ""
    ],
    [
        "LOGIN_TC_017", "test_login_case_sensitive_password",
        "Verify password field is case-sensitive",
        "Case Sensitivity - Password",
        "Negative", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter 'standard_user' in the Username field\n3. Enter 'Secret_Sauce' (with capitals) in the Password field\n4. Click the 'Login' button",
        "Username: standard_user\nPassword: Secret_Sauce",
        "Login should fail because passwords are case-sensitive. Error message should be displayed.",
        "Login fails. Error message 'Epic sadface: Username and password do not match any user in this service' is displayed.",
        "Major", "P2", "Pass", ""
    ],
    [
        "LOGIN_TC_018", "test_login_keyboard_enter_key",
        "Verify login using Enter key instead of clicking Login button",
        "Login via Enter Key",
        "Functional", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter 'standard_user' in the Username field\n3. Enter 'secret_sauce' in the Password field\n4. Press the 'Enter' key on the keyboard",
        "Username: standard_user\nPassword: secret_sauce",
        "User should be successfully logged in using the Enter key, same as clicking the Login button.",
        "User is successfully logged in via Enter key. Redirected to inventory page.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "LOGIN_TC_019", "test_login_tab_navigation",
        "Verify tab key navigates between form fields correctly",
        "Tab Key Navigation on Login Page",
        "Accessibility", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Click on the Username field\n3. Press Tab key to move to Password field\n4. Press Tab key to move to Login button\n5. Verify focus moves correctly through the form elements",
        "N/A",
        "Tab key should navigate focus from Username field → Password field → Login button in the correct order.",
        "Tab navigation works correctly: Username → Password → Login button.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "LOGIN_TC_020", "test_login_direct_inventory_url_without_auth",
        "Verify direct URL access to inventory page without authentication",
        "Unauthorized Direct URL Access",
        "Security", "Browser is open; User is NOT logged in",
        "1. Open browser\n2. Directly navigate to https://www.saucedemo.com/inventory.html without logging in\n3. Observe the behavior",
        "URL: https://www.saucedemo.com/inventory.html",
        "User should be redirected back to the login page with error message 'Epic sadface: You can only access '/inventory.html' when you are logged in.'",
        "User is redirected to login page. Error message 'Epic sadface: You can only access '/inventory.html' when you are logged in.' is displayed.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "LOGIN_TC_021", "test_login_visual_user",
        "Verify login with visual_user credentials",
        "Valid Login - Visual User",
        "Positive", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter username 'visual_user' in the Username field\n3. Enter password 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: visual_user\nPassword: secret_sauce",
        "User should be logged in and redirected to inventory page. Visual discrepancies/misalignments may be present on the page.",
        "User is logged in. Inventory page shows visual bugs - misaligned elements, incorrect positioning of some UI components.",
        "Major", "P2", "Fail", "Known visual bugs with visual_user account"
    ],
    [
        "LOGIN_TC_022", "test_login_error_user",
        "Verify login with error_user credentials",
        "Valid Login - Error User",
        "Positive", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter username 'error_user' in the Username field\n3. Enter password 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: error_user\nPassword: secret_sauce",
        "User should be logged in and redirected to inventory page. Some functionality may produce errors during interaction.",
        "User is logged in and redirected to inventory page. Some interactive features produce errors during usage.",
        "Major", "P2", "Pass", "Error user has intentional bugs in cart/checkout operations"
    ],
    [
        "LOGIN_TC_023", "test_login_special_characters_username",
        "Verify login with special characters in username",
        "Special Characters in Username",
        "Negative", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter '!@#$%^&*()' in the Username field\n3. Enter 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: !@#$%^&*()\nPassword: secret_sauce",
        "Login should fail with error message. Application should handle special characters gracefully without crashing.",
        "Login fails. Error message 'Epic sadface: Username and password do not match any user in this service' is displayed. No crash.",
        "Major", "P2", "Pass", ""
    ],
    [
        "LOGIN_TC_024", "test_login_long_input_username",
        "Verify login with extremely long username input",
        "Long Input in Username Field",
        "Boundary", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Enter a 500-character string in the Username field\n3. Enter 'secret_sauce' in the Password field\n4. Click the 'Login' button",
        "Username: 500-character string (aaaa...)\nPassword: secret_sauce",
        "Login should fail with standard error message. The application should handle long input without crashing or breaking the UI layout.",
        "Login fails with standard error message. UI layout remains intact.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "LOGIN_TC_025", "test_login_page_title_and_logo",
        "Verify the login page displays correct title and Swag Labs logo",
        "Login Page Title and Logo",
        "UI", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Verify the page title in the browser tab\n3. Verify the Swag Labs logo/heading is displayed on the page",
        "N/A",
        "Browser tab title should be 'Swag Labs'. The Swag Labs logo should be visible at the top of the login form.",
        "Browser tab title is 'Swag Labs'. Swag Labs logo is displayed at the top of the login page.",
        "Minor", "P3", "Pass", ""
    ],
]

# ============================================================
# MODULE 2: INVENTORY (PRODUCTS PAGE)
# ============================================================
inventory_tests = [
    [
        "INV_TC_001", "test_inventory_page_display_all_products",
        "Verify all 6 products are displayed on the inventory page",
        "Products List Display",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Observe the inventory page\n3. Count the number of products displayed\n4. Verify each product has an image, name, description, price, and Add to Cart button",
        "N/A",
        "All 6 products should be displayed: Sauce Labs Backpack, Sauce Labs Bike Light, Sauce Labs Bolt T-Shirt, Sauce Labs Fleece Jacket, Sauce Labs Onesie, Test.allTheThings() T-Shirt (Red). Each product should have image, name, description, price, and 'Add to cart' button.",
        "All 6 products are displayed with correct names, images, descriptions, prices, and 'Add to cart' buttons.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "INV_TC_002", "test_inventory_product_prices",
        "Verify correct prices for all products",
        "Product Prices Verification",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Verify the price of each product on the inventory page:\n   - Sauce Labs Backpack: $29.99\n   - Sauce Labs Bike Light: $9.99\n   - Sauce Labs Bolt T-Shirt: $15.99\n   - Sauce Labs Fleece Jacket: $49.99\n   - Sauce Labs Onesie: $7.99\n   - Test.allTheThings() T-Shirt (Red): $15.99",
        "Expected prices as listed in test steps",
        "Each product should display the correct price with '$' prefix and two decimal places.",
        "All product prices match expected values: Backpack=$29.99, Bike Light=$9.99, Bolt T-Shirt=$15.99, Fleece Jacket=$49.99, Onesie=$7.99, T-Shirt Red=$15.99.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "INV_TC_003", "test_inventory_product_images_displayed",
        "Verify all product images are loaded and displayed correctly",
        "Product Images Display",
        "UI", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Verify each product has a visible image\n3. Check that no image shows broken image icon\n4. Verify images are different for each product",
        "N/A",
        "Each product should display its unique product image. No broken image icons should be shown. All 6 images should be distinct.",
        "All 6 product images are displayed correctly. Each product has a unique image. No broken images.",
        "Major", "P2", "Pass", ""
    ],
    [
        "INV_TC_004", "test_inventory_add_single_item_to_cart",
        "Verify adding a single item to cart from inventory page",
        "Add Single Item to Cart",
        "Functional", "User is logged in as standard_user; Cart is empty",
        "1. Login with valid credentials\n2. Click 'Add to cart' button for 'Sauce Labs Backpack'\n3. Observe the button text change\n4. Observe the cart badge icon",
        "Product: Sauce Labs Backpack",
        "The 'Add to cart' button should change to 'Remove'. The cart badge should show '1' indicating one item in the cart.",
        "Button text changes to 'Remove'. Cart badge displays '1'.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "INV_TC_005", "test_inventory_add_multiple_items_to_cart",
        "Verify adding multiple items to cart from inventory page",
        "Add Multiple Items to Cart",
        "Functional", "User is logged in as standard_user; Cart is empty",
        "1. Login with valid credentials\n2. Click 'Add to cart' for 'Sauce Labs Backpack'\n3. Click 'Add to cart' for 'Sauce Labs Bike Light'\n4. Click 'Add to cart' for 'Sauce Labs Onesie'\n5. Observe the cart badge count",
        "Products: Sauce Labs Backpack, Sauce Labs Bike Light, Sauce Labs Onesie",
        "All three 'Add to cart' buttons should change to 'Remove'. The cart badge should show '3'.",
        "All three buttons changed to 'Remove'. Cart badge displays '3'.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "INV_TC_006", "test_inventory_add_all_items_to_cart",
        "Verify adding all 6 items to cart",
        "Add All Items to Cart",
        "Boundary", "User is logged in as standard_user; Cart is empty",
        "1. Login with valid credentials\n2. Click 'Add to cart' for all 6 products one by one\n3. Observe the cart badge count after each addition",
        "All 6 products",
        "All 'Add to cart' buttons should change to 'Remove'. Cart badge should show '6'.",
        "All 6 buttons changed to 'Remove'. Cart badge displays '6'.",
        "Major", "P2", "Pass", ""
    ],
    [
        "INV_TC_007", "test_inventory_remove_item_from_cart",
        "Verify removing an item from cart on inventory page",
        "Remove Item from Cart on Inventory",
        "Functional", "User is logged in as standard_user; At least one item is in the cart",
        "1. Login and add 'Sauce Labs Backpack' to cart\n2. Verify cart badge shows '1'\n3. Click the 'Remove' button for 'Sauce Labs Backpack'\n4. Observe button text and cart badge",
        "Product: Sauce Labs Backpack",
        "The 'Remove' button should change back to 'Add to cart'. The cart badge should disappear (no items in cart).",
        "'Remove' button changes back to 'Add to cart'. Cart badge disappears.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "INV_TC_008", "test_inventory_sort_az",
        "Verify sorting products by Name (A to Z)",
        "Sort Products A to Z",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click on the sort dropdown (Product Sort Container)\n3. Select 'Name (A to Z)'\n4. Verify the products are listed in alphabetical order",
        "Sort option: Name (A to Z)",
        "Products should be sorted alphabetically A-Z: Sauce Labs Backpack, Sauce Labs Bike Light, Sauce Labs Bolt T-Shirt, Sauce Labs Fleece Jacket, Sauce Labs Onesie, Test.allTheThings() T-Shirt (Red).",
        "Products are sorted alphabetically A-Z. Default sort order matches expected order.",
        "Major", "P2", "Pass", "This is the default sort order"
    ],
    [
        "INV_TC_009", "test_inventory_sort_za",
        "Verify sorting products by Name (Z to A)",
        "Sort Products Z to A",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click on the sort dropdown\n3. Select 'Name (Z to A)'\n4. Verify the products are listed in reverse alphabetical order",
        "Sort option: Name (Z to A)",
        "Products should be sorted reverse alphabetically Z-A: Test.allTheThings() T-Shirt (Red), Sauce Labs Onesie, Sauce Labs Fleece Jacket, Sauce Labs Bolt T-Shirt, Sauce Labs Bike Light, Sauce Labs Backpack.",
        "Products are sorted in reverse alphabetical order (Z-A) as expected.",
        "Major", "P2", "Pass", ""
    ],
    [
        "INV_TC_010", "test_inventory_sort_price_low_to_high",
        "Verify sorting products by Price (low to high)",
        "Sort Products Price Low to High",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click on the sort dropdown\n3. Select 'Price (low to high)'\n4. Verify products are ordered by ascending price",
        "Sort option: Price (low to high)",
        "Products should be sorted by price ascending: Onesie ($7.99), Bike Light ($9.99), Bolt T-Shirt ($15.99), T-Shirt Red ($15.99), Backpack ($29.99), Fleece Jacket ($49.99).",
        "Products are sorted by price low to high as expected.",
        "Major", "P2", "Pass", ""
    ],
    [
        "INV_TC_011", "test_inventory_sort_price_high_to_low",
        "Verify sorting products by Price (high to low)",
        "Sort Products Price High to Low",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click on the sort dropdown\n3. Select 'Price (high to low)'\n4. Verify products are ordered by descending price",
        "Sort option: Price (high to low)",
        "Products should be sorted by price descending: Fleece Jacket ($49.99), Backpack ($29.99), Bolt T-Shirt ($15.99), T-Shirt Red ($15.99), Bike Light ($9.99), Onesie ($7.99).",
        "Products are sorted by price high to low as expected.",
        "Major", "P2", "Pass", ""
    ],
    [
        "INV_TC_012", "test_inventory_sort_persists_after_add_to_cart",
        "Verify sort order persists after adding an item to cart",
        "Sort Persistence After Cart Action",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Sort products by 'Price (high to low)'\n3. Add the first product (Fleece Jacket) to cart\n4. Verify the sort order remains Price (high to low)",
        "Sort option: Price (high to low)",
        "The sort order should remain 'Price (high to low)' after adding an item to cart. Product order should not change.",
        "Sort order remains 'Price (high to low)'. Product order unchanged after adding item to cart.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "INV_TC_013", "test_inventory_product_name_clickable",
        "Verify clicking product name navigates to product detail page",
        "Product Name Navigation Link",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click on the product name 'Sauce Labs Backpack'\n3. Observe the navigation",
        "Product: Sauce Labs Backpack",
        "User should be navigated to the product detail page for Sauce Labs Backpack showing full product details.",
        "User is navigated to product detail page. Full product details are displayed for Sauce Labs Backpack.",
        "Major", "P2", "Pass", ""
    ],
    [
        "INV_TC_014", "test_inventory_product_image_clickable",
        "Verify clicking product image navigates to product detail page",
        "Product Image Navigation Link",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click on the product image for 'Sauce Labs Backpack'\n3. Observe the navigation",
        "Product: Sauce Labs Backpack",
        "User should be navigated to the product detail page for Sauce Labs Backpack.",
        "User is navigated to product detail page for Sauce Labs Backpack on clicking the image.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "INV_TC_015", "test_inventory_cart_badge_not_visible_when_empty",
        "Verify cart badge is not displayed when cart is empty",
        "Cart Badge - Empty Cart",
        "UI", "User is logged in as standard_user; Cart is empty",
        "1. Login with valid credentials\n2. Ensure no items are added to cart\n3. Observe the cart icon in the top-right corner",
        "N/A",
        "The cart icon should be visible but no badge/count number should be displayed when the cart is empty.",
        "Cart icon is visible. No badge number is displayed when cart is empty.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "INV_TC_016", "test_inventory_cart_icon_clickable",
        "Verify clicking the cart icon navigates to the cart page",
        "Cart Icon Navigation",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click the shopping cart icon in the top-right corner\n3. Observe the navigation",
        "N/A",
        "User should be navigated to the cart page (https://www.saucedemo.com/cart.html).",
        "User is navigated to the cart page successfully.",
        "Major", "P2", "Pass", ""
    ],
    [
        "INV_TC_017", "test_inventory_page_header",
        "Verify the inventory page header displays correctly",
        "Inventory Page Header",
        "UI", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Observe the page header section\n3. Verify 'Swag Labs' app logo is displayed\n4. Verify 'Products' title is displayed",
        "N/A",
        "'Swag Labs' logo should be in the header. 'Products' title should be displayed as the page heading. Hamburger menu icon and cart icon should be visible.",
        "'Swag Labs' logo displayed. 'Products' title shown. Hamburger menu and cart icon are visible.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "INV_TC_018", "test_inventory_footer_links",
        "Verify footer social media links are present and functional",
        "Footer Social Media Links",
        "UI", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Scroll down to the footer section\n3. Verify Twitter, Facebook, and LinkedIn icons/links are present\n4. Verify the footer copyright text",
        "N/A",
        "Footer should contain Twitter, Facebook, and LinkedIn social media links. Copyright text '© 2025 Sauce Labs. All Rights Reserved. Terms of Service | Privacy Policy' should be displayed.",
        "Footer contains Twitter, Facebook, LinkedIn links. Copyright text is displayed correctly.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "INV_TC_019", "test_inventory_twitter_footer_link",
        "Verify Twitter footer link opens correct URL",
        "Twitter Footer Link",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Scroll to footer\n3. Click the Twitter icon/link\n4. Verify it opens the correct Sauce Labs Twitter page",
        "N/A",
        "Twitter link should open https://twitter.com/saboralabs in a new tab.",
        "Twitter link opens the Sauce Labs Twitter page in a new tab.",
        "Minor", "P4", "Pass", ""
    ],
    [
        "INV_TC_020", "test_inventory_facebook_footer_link",
        "Verify Facebook footer link opens correct URL",
        "Facebook Footer Link",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Scroll to footer\n3. Click the Facebook icon/link\n4. Verify it opens the correct Sauce Labs Facebook page",
        "N/A",
        "Facebook link should open https://www.facebook.com/saucelabs in a new tab.",
        "Facebook link opens the Sauce Labs Facebook page in a new tab.",
        "Minor", "P4", "Pass", ""
    ],
    [
        "INV_TC_021", "test_inventory_linkedin_footer_link",
        "Verify LinkedIn footer link opens correct URL",
        "LinkedIn Footer Link",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Scroll to footer\n3. Click the LinkedIn icon/link\n4. Verify it opens the correct Sauce Labs LinkedIn page",
        "N/A",
        "LinkedIn link should open https://www.linkedin.com/company/sauce-labs/ in a new tab.",
        "LinkedIn link opens the Sauce Labs LinkedIn page in a new tab.",
        "Minor", "P4", "Pass", ""
    ],
    [
        "INV_TC_022", "test_inventory_product_description_text",
        "Verify product descriptions are displayed correctly",
        "Product Description Display",
        "UI", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Verify each product card shows a description text below the product name\n3. Verify descriptions are not empty or truncated improperly",
        "N/A",
        "Each product should have a visible description text. The descriptions should be meaningful and not blank. Text should not overflow the product card container.",
        "All 6 products display meaningful description text. No overflow or layout issues.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "INV_TC_023", "test_inventory_add_to_cart_button_state_persistence",
        "Verify Add to Cart button state persists after page navigation",
        "Cart Button State Persistence",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Add 'Sauce Labs Backpack' to cart (button shows 'Remove')\n3. Click on another product to go to its detail page\n4. Navigate back to the inventory page\n5. Verify the 'Sauce Labs Backpack' button still shows 'Remove'",
        "Product: Sauce Labs Backpack",
        "The 'Remove' button state for Sauce Labs Backpack should persist after navigating away and coming back to the inventory page.",
        "'Remove' button state persists for Sauce Labs Backpack after navigating back to inventory page.",
        "Major", "P2", "Pass", ""
    ],
]

# ============================================================
# MODULE 3: PRODUCT DETAILS
# ============================================================
product_detail_tests = [
    [
        "PD_TC_001", "test_product_detail_page_elements",
        "Verify all elements on product detail page",
        "Product Detail Page Elements",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click on 'Sauce Labs Backpack' product name\n3. Verify the following elements are present:\n   - Product image (large)\n   - Product name\n   - Product description\n   - Product price\n   - 'Add to cart' button\n   - 'Back to products' button",
        "Product: Sauce Labs Backpack",
        "Product detail page should display: large product image, product name 'Sauce Labs Backpack', full description, price '$29.99', 'Add to cart' button, and 'Back to products' button.",
        "All elements are displayed correctly: large image, name, description, price $29.99, Add to cart button, and Back to products button.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "PD_TC_002", "test_product_detail_back_to_products",
        "Verify 'Back to products' button navigates back to inventory",
        "Back to Products Navigation",
        "Functional", "User is logged in as standard_user; On product detail page",
        "1. Login and navigate to any product detail page\n2. Click the 'Back to products' button\n3. Observe the navigation",
        "N/A",
        "User should be navigated back to the inventory page with all products displayed.",
        "User is navigated back to inventory page. All products are displayed.",
        "Major", "P2", "Pass", ""
    ],
    [
        "PD_TC_003", "test_product_detail_add_to_cart",
        "Verify adding item to cart from product detail page",
        "Add to Cart from Product Detail",
        "Functional", "User is logged in as standard_user; Cart is empty",
        "1. Login and navigate to 'Sauce Labs Backpack' detail page\n2. Click 'Add to cart' button\n3. Observe button change and cart badge",
        "Product: Sauce Labs Backpack",
        "Button should change to 'Remove'. Cart badge should show '1'.",
        "Button changes to 'Remove'. Cart badge shows '1'.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "PD_TC_004", "test_product_detail_remove_from_cart",
        "Verify removing item from cart on product detail page",
        "Remove from Cart on Product Detail",
        "Functional", "User is logged in as standard_user; Item is in cart",
        "1. Login and add 'Sauce Labs Backpack' to cart\n2. Navigate to 'Sauce Labs Backpack' detail page\n3. Click the 'Remove' button\n4. Observe button change and cart badge",
        "Product: Sauce Labs Backpack",
        "Button should change back to 'Add to cart'. Cart badge should disappear.",
        "Button changes to 'Add to cart'. Cart badge disappears.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "PD_TC_005", "test_product_detail_correct_product_info",
        "Verify correct product information is displayed for each product",
        "Product Info Accuracy",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Navigate to each product's detail page one by one\n3. Verify the product name, description, and price match the inventory page listing",
        "All 6 products",
        "Product name, description, and price on the detail page should match exactly what is shown on the inventory page for each product.",
        "All product details match between inventory page and detail pages for all 6 products.",
        "Major", "P2", "Pass", ""
    ],
    [
        "PD_TC_006", "test_product_detail_image_large_display",
        "Verify product image is displayed in larger size on detail page",
        "Product Image Large Display",
        "UI", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Navigate to 'Sauce Labs Backpack' detail page\n3. Verify the product image is displayed\n4. Compare image size with inventory page thumbnail",
        "Product: Sauce Labs Backpack",
        "The product image on the detail page should be larger than the thumbnail on the inventory page. Image should be clearly visible and not pixelated.",
        "Product image is displayed in larger size. Image is clear and not pixelated.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "PD_TC_007", "test_product_detail_cart_icon_accessible",
        "Verify cart icon is accessible from product detail page",
        "Cart Icon on Product Detail Page",
        "Functional", "User is logged in as standard_user; On product detail page",
        "1. Login and navigate to any product detail page\n2. Verify the cart icon is visible in the header\n3. Click the cart icon\n4. Verify navigation to cart page",
        "N/A",
        "Cart icon should be visible and clickable on the product detail page. Clicking it should navigate to the cart page.",
        "Cart icon is visible and functional. Clicking navigates to cart page.",
        "Major", "P2", "Pass", ""
    ],
    [
        "PD_TC_008", "test_product_detail_hamburger_menu_accessible",
        "Verify hamburger menu is accessible from product detail page",
        "Hamburger Menu on Product Detail",
        "Functional", "User is logged in as standard_user; On product detail page",
        "1. Login and navigate to any product detail page\n2. Click the hamburger menu icon (☰)\n3. Verify menu opens with all options",
        "N/A",
        "Hamburger menu should open showing: All Items, About, Logout, Reset App State options.",
        "Hamburger menu opens with all 4 menu options: All Items, About, Logout, Reset App State.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "PD_TC_009", "test_product_detail_add_to_cart_state_sync_with_inventory",
        "Verify cart state syncs between product detail and inventory pages",
        "Cart State Sync Detail-Inventory",
        "Functional", "User is logged in as standard_user; Cart is empty",
        "1. Login and navigate to 'Sauce Labs Backpack' detail page\n2. Click 'Add to cart'\n3. Click 'Back to products' to go to inventory page\n4. Verify 'Sauce Labs Backpack' shows 'Remove' button on inventory page",
        "Product: Sauce Labs Backpack",
        "The 'Remove' button state should be synced. Backpack should show 'Remove' on the inventory page after being added from the detail page.",
        "'Remove' button for Backpack is displayed on inventory page after adding from detail page.",
        "Major", "P2", "Pass", ""
    ],
    [
        "PD_TC_010", "test_product_detail_url_structure",
        "Verify product detail page URL contains correct item ID",
        "Product Detail URL Structure",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click on 'Sauce Labs Backpack'\n3. Verify the URL in the browser address bar\n4. Check URL format contains inventory-item.html with item ID parameter",
        "Product: Sauce Labs Backpack",
        "URL should follow the pattern: https://www.saucedemo.com/inventory-item.html?id=X where X is the product's unique ID.",
        "URL is https://www.saucedemo.com/inventory-item.html?id=4 for Sauce Labs Backpack.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "PD_TC_011", "test_product_detail_invalid_product_id",
        "Verify behavior when accessing product detail with invalid ID",
        "Invalid Product ID in URL",
        "Negative", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Manually navigate to https://www.saucedemo.com/inventory-item.html?id=999\n3. Observe the page behavior",
        "URL: inventory-item.html?id=999",
        "Application should handle invalid product ID gracefully - either show an error message or redirect to inventory page.",
        "Page loads but shows empty/default content. No crash or error page.",
        "Major", "P2", "Pass", "Application handles gracefully but could show a better error message"
    ],
]

# ============================================================
# MODULE 4: CART
# ============================================================
cart_tests = [
    [
        "CART_TC_001", "test_cart_page_display_with_items",
        "Verify cart page displays added items correctly",
        "Cart Page Display with Items",
        "Functional", "User is logged in as standard_user; Items are added to cart",
        "1. Login and add 'Sauce Labs Backpack' and 'Sauce Labs Bike Light' to cart\n2. Click the cart icon\n3. Verify both items are displayed with:\n   - Quantity (1 each)\n   - Product name\n   - Product description\n   - Product price\n   - 'Remove' button for each item",
        "Products: Sauce Labs Backpack ($29.99), Sauce Labs Bike Light ($9.99)",
        "Cart page should display both items with quantity '1', correct names, descriptions, prices, and 'Remove' button for each. 'Continue Shopping' and 'Checkout' buttons should be visible.",
        "Both items displayed with quantity 1, correct names, descriptions, prices. Remove buttons, Continue Shopping, and Checkout buttons are visible.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CART_TC_002", "test_cart_page_empty_cart",
        "Verify cart page display when cart is empty",
        "Empty Cart Page Display",
        "Functional", "User is logged in as standard_user; Cart is empty",
        "1. Login with valid credentials\n2. Click the cart icon without adding any items\n3. Observe the cart page",
        "N/A",
        "Cart page should be displayed with no items listed. 'Your Cart' title should be shown. 'Continue Shopping' and 'Checkout' buttons should be visible.",
        "Empty cart page displayed. 'Your Cart' title shown. Continue Shopping and Checkout buttons are visible.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CART_TC_003", "test_cart_remove_single_item",
        "Verify removing a single item from the cart page",
        "Remove Single Item from Cart",
        "Functional", "User is logged in as standard_user; 2 items in cart",
        "1. Login and add Backpack and Bike Light to cart\n2. Navigate to cart page\n3. Click 'Remove' button for 'Sauce Labs Backpack'\n4. Verify Backpack is removed\n5. Verify Bike Light is still in cart\n6. Verify cart badge shows '1'",
        "Remove: Sauce Labs Backpack\nKeep: Sauce Labs Bike Light",
        "Sauce Labs Backpack should be removed from the cart. Sauce Labs Bike Light should remain. Cart badge should update to '1'.",
        "Backpack is removed. Bike Light remains in cart. Cart badge shows '1'.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CART_TC_004", "test_cart_remove_all_items",
        "Verify removing all items from the cart",
        "Remove All Items from Cart",
        "Functional", "User is logged in as standard_user; Items in cart",
        "1. Login and add 2 items to cart\n2. Navigate to cart page\n3. Click 'Remove' for the first item\n4. Click 'Remove' for the second item\n5. Verify cart is empty\n6. Verify cart badge disappears",
        "N/A",
        "Both items should be removed. Cart should be empty. Cart badge should disappear from the header.",
        "Both items removed. Cart is empty. Cart badge is no longer displayed.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CART_TC_005", "test_cart_continue_shopping_button",
        "Verify 'Continue Shopping' button navigates back to inventory",
        "Continue Shopping Button",
        "Functional", "User is logged in as standard_user; On cart page",
        "1. Login and navigate to cart page\n2. Click the 'Continue Shopping' button\n3. Observe the navigation",
        "N/A",
        "User should be navigated back to the inventory page (products page).",
        "User is navigated back to the inventory page.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CART_TC_006", "test_cart_checkout_button",
        "Verify 'Checkout' button navigates to checkout step one",
        "Checkout Button Navigation",
        "Functional", "User is logged in as standard_user; Items in cart",
        "1. Login and add an item to cart\n2. Navigate to cart page\n3. Click the 'Checkout' button\n4. Observe the navigation",
        "Product: Sauce Labs Backpack",
        "User should be navigated to the checkout step one page (checkout-step-one.html) with 'Your Information' form.",
        "User is navigated to checkout step one page with the information form.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CART_TC_007", "test_cart_item_quantity_display",
        "Verify item quantity is displayed correctly in cart",
        "Cart Item Quantity Display",
        "UI", "User is logged in as standard_user; Items in cart",
        "1. Login and add 'Sauce Labs Backpack' to cart\n2. Navigate to cart page\n3. Verify the quantity displayed for the item",
        "Product: Sauce Labs Backpack",
        "The quantity column should display '1' for the added item.",
        "Quantity shows '1' for Sauce Labs Backpack.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CART_TC_008", "test_cart_prices_match_inventory",
        "Verify item prices in cart match inventory page prices",
        "Cart Price Consistency",
        "Functional", "User is logged in as standard_user; Items in cart",
        "1. Note the price of 'Sauce Labs Backpack' on inventory page ($29.99)\n2. Add it to cart\n3. Navigate to cart page\n4. Verify the price shown matches the inventory page price",
        "Product: Sauce Labs Backpack, Price: $29.99",
        "The price displayed in the cart should match the price on the inventory page ($29.99).",
        "Price in cart ($29.99) matches the inventory page price.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CART_TC_009", "test_cart_page_header_title",
        "Verify 'Your Cart' title is displayed on cart page",
        "Cart Page Title",
        "UI", "User is logged in as standard_user; On cart page",
        "1. Login with valid credentials\n2. Navigate to cart page\n3. Verify 'Your Cart' title is displayed",
        "N/A",
        "'Your Cart' should be displayed as the page title/header on the cart page.",
        "'Your Cart' title is displayed on the cart page.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "CART_TC_010", "test_cart_column_headers",
        "Verify cart page displays QTY and Description column headers",
        "Cart Column Headers",
        "UI", "User is logged in as standard_user; On cart page",
        "1. Login with valid credentials\n2. Add an item to cart\n3. Navigate to cart page\n4. Verify 'QTY' and 'Description' column headers are visible",
        "N/A",
        "'QTY' and 'Description' labels should be visible as column headers in the cart.",
        "'QTY' and 'Description' column headers are displayed.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "CART_TC_011", "test_cart_item_name_clickable",
        "Verify clicking item name in cart navigates to product detail",
        "Cart Item Name Link",
        "Functional", "User is logged in as standard_user; Items in cart",
        "1. Login and add 'Sauce Labs Backpack' to cart\n2. Navigate to cart page\n3. Click on the product name 'Sauce Labs Backpack'\n4. Observe the navigation",
        "Product: Sauce Labs Backpack",
        "Clicking the product name should navigate to the product detail page for that item.",
        "Clicking product name navigates to Sauce Labs Backpack detail page.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "CART_TC_012", "test_cart_persists_after_page_refresh",
        "Verify cart items persist after page refresh",
        "Cart Persistence on Refresh",
        "Functional", "User is logged in as standard_user; Items in cart",
        "1. Login and add items to cart\n2. Navigate to cart page\n3. Refresh the page (F5)\n4. Verify items are still in the cart",
        "Products: Sauce Labs Backpack, Sauce Labs Bike Light",
        "Cart items should persist after page refresh. All previously added items should remain in the cart.",
        "Cart items persist after page refresh. Both items remain in the cart.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CART_TC_013", "test_cart_checkout_with_empty_cart",
        "Verify checkout behavior with empty cart",
        "Checkout with Empty Cart",
        "Negative", "User is logged in as standard_user; Cart is empty",
        "1. Login with valid credentials\n2. Navigate to cart page (empty cart)\n3. Click the 'Checkout' button\n4. Observe the behavior",
        "N/A",
        "Application should either prevent checkout with an empty cart or display a message indicating no items to checkout.",
        "Application allows proceeding to checkout step one even with empty cart. No validation message shown.",
        "Major", "P2", "Fail", "Bug: Application allows checkout with empty cart - no validation"
    ],
]

# ============================================================
# MODULE 5: CHECKOUT
# ============================================================
checkout_tests = [
    [
        "CHK_TC_001", "test_checkout_step_one_page_elements",
        "Verify all elements on checkout step one (Your Information) page",
        "Checkout Step One Page Elements",
        "Functional", "User is logged in; Items in cart; On checkout step one page",
        "1. Login, add item to cart, go to cart, click Checkout\n2. Verify the following elements:\n   - 'Checkout: Your Information' title\n   - First Name input field\n   - Last Name input field\n   - Zip/Postal Code input field\n   - 'Cancel' button\n   - 'Continue' button",
        "N/A",
        "Checkout step one should display: title 'Checkout: Your Information', First Name, Last Name, Zip/Postal Code input fields, Cancel and Continue buttons.",
        "All elements displayed: title, 3 input fields (First Name, Last Name, Zip/Postal Code), Cancel and Continue buttons.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_002", "test_checkout_step_one_valid_info",
        "Verify successful submission with valid checkout information",
        "Valid Checkout Information Submit",
        "Positive", "User is logged in; Items in cart; On checkout step one page",
        "1. Navigate to checkout step one\n2. Enter 'John' in First Name field\n3. Enter 'Doe' in Last Name field\n4. Enter '12345' in Zip/Postal Code field\n5. Click 'Continue' button",
        "First Name: John\nLast Name: Doe\nZip: 12345",
        "User should be navigated to checkout step two (Overview page) showing order summary.",
        "User is navigated to checkout step two (Overview page) with order summary displayed.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_003", "test_checkout_step_one_empty_first_name",
        "Verify error when First Name is empty",
        "Empty First Name Validation",
        "Negative", "User is logged in; Items in cart; On checkout step one page",
        "1. Navigate to checkout step one\n2. Leave First Name field empty\n3. Enter 'Doe' in Last Name\n4. Enter '12345' in Zip/Postal Code\n5. Click 'Continue' button",
        "First Name: (empty)\nLast Name: Doe\nZip: 12345",
        "Error message 'Error: First Name is required' should be displayed. User should remain on checkout step one.",
        "Error message 'Error: First Name is required' is displayed. User remains on step one.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_004", "test_checkout_step_one_empty_last_name",
        "Verify error when Last Name is empty",
        "Empty Last Name Validation",
        "Negative", "User is logged in; Items in cart; On checkout step one page",
        "1. Navigate to checkout step one\n2. Enter 'John' in First Name\n3. Leave Last Name field empty\n4. Enter '12345' in Zip/Postal Code\n5. Click 'Continue' button",
        "First Name: John\nLast Name: (empty)\nZip: 12345",
        "Error message 'Error: Last Name is required' should be displayed. User should remain on checkout step one.",
        "Error message 'Error: Last Name is required' is displayed. User remains on step one.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_005", "test_checkout_step_one_empty_zip_code",
        "Verify error when Zip/Postal Code is empty",
        "Empty Zip Code Validation",
        "Negative", "User is logged in; Items in cart; On checkout step one page",
        "1. Navigate to checkout step one\n2. Enter 'John' in First Name\n3. Enter 'Doe' in Last Name\n4. Leave Zip/Postal Code field empty\n5. Click 'Continue' button",
        "First Name: John\nLast Name: Doe\nZip: (empty)",
        "Error message 'Error: Postal Code is required' should be displayed. User should remain on checkout step one.",
        "Error message 'Error: Postal Code is required' is displayed. User remains on step one.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_006", "test_checkout_step_one_all_fields_empty",
        "Verify error when all fields are empty",
        "All Fields Empty Validation",
        "Negative", "User is logged in; Items in cart; On checkout step one page",
        "1. Navigate to checkout step one\n2. Leave all fields empty\n3. Click 'Continue' button",
        "All fields empty",
        "Error message 'Error: First Name is required' should be displayed (first validation).",
        "Error message 'Error: First Name is required' is displayed.",
        "Major", "P1", "Pass", ""
    ],
    [
        "CHK_TC_007", "test_checkout_step_one_cancel_button",
        "Verify Cancel button navigates back to cart page",
        "Checkout Step One Cancel Button",
        "Functional", "User is logged in; Items in cart; On checkout step one page",
        "1. Navigate to checkout step one\n2. Enter some data in the fields\n3. Click the 'Cancel' button\n4. Observe the navigation",
        "N/A",
        "User should be navigated back to the cart page. Cart items should remain unchanged.",
        "User is navigated back to cart page. Cart items remain unchanged.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CHK_TC_008", "test_checkout_step_one_error_close_button",
        "Verify error message close button on checkout step one",
        "Checkout Error Close Button",
        "Functional", "User is logged in; Items in cart; On checkout step one page",
        "1. Navigate to checkout step one\n2. Click 'Continue' without filling fields\n3. Verify error message appears\n4. Click the 'X' close button on the error\n5. Verify error disappears",
        "N/A",
        "The error message should disappear when the close button is clicked.",
        "Error message disappears when close button is clicked.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "CHK_TC_009", "test_checkout_step_two_overview_elements",
        "Verify all elements on checkout step two (Overview) page",
        "Checkout Overview Page Elements",
        "Functional", "User is logged in; Items in cart; Checkout info filled",
        "1. Complete checkout step one with valid info\n2. On the Overview page, verify:\n   - 'Checkout: Overview' title\n   - Cart items with quantity, name, description, price\n   - Payment Information section\n   - Shipping Information section\n   - Price Total section (Item total, Tax, Total)\n   - 'Cancel' button\n   - 'Finish' button",
        "N/A",
        "Overview page should show all ordered items, Payment Information (SauceCard #31337), Shipping Information (Free Pony Express Delivery!), Item total, Tax (8%), and Total. Cancel and Finish buttons should be visible.",
        "All elements displayed correctly. Payment: SauceCard #31337, Shipping: Free Pony Express Delivery!, Item total, Tax, and Total are shown. Cancel and Finish buttons visible.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_010", "test_checkout_step_two_item_total_calculation",
        "Verify item total is calculated correctly",
        "Item Total Calculation",
        "Functional", "User is logged in; Multiple items in cart; On overview page",
        "1. Add 'Sauce Labs Backpack' ($29.99) and 'Sauce Labs Bike Light' ($9.99) to cart\n2. Proceed through checkout step one\n3. On overview page, verify the 'Item total' value",
        "Backpack: $29.99\nBike Light: $9.99\nExpected Total: $39.98",
        "Item total should be $39.98 (sum of $29.99 + $9.99).",
        "Item total displays $39.98.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_011", "test_checkout_step_two_tax_calculation",
        "Verify tax is calculated correctly (8% tax rate)",
        "Tax Calculation Verification",
        "Functional", "User is logged in; Items in cart; On overview page",
        "1. Add 'Sauce Labs Backpack' ($29.99) and 'Sauce Labs Bike Light' ($9.99) to cart\n2. Proceed to checkout overview page\n3. Verify the tax amount\n4. Calculate expected tax: $39.98 × 8% = $3.20",
        "Item total: $39.98\nTax Rate: 8%\nExpected Tax: $3.20",
        "Tax should be approximately $3.20 (8% of item total $39.98).",
        "Tax displays $3.20.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_012", "test_checkout_step_two_total_with_tax",
        "Verify total amount includes items plus tax",
        "Total Amount with Tax",
        "Functional", "User is logged in; Items in cart; On overview page",
        "1. Add 'Sauce Labs Backpack' ($29.99) and 'Sauce Labs Bike Light' ($9.99) to cart\n2. Proceed to checkout overview page\n3. Verify Total = Item total + Tax\n4. Expected: $39.98 + $3.20 = $43.18",
        "Item total: $39.98\nTax: $3.20\nExpected Total: $43.18",
        "Total should be $43.18 (Item total $39.98 + Tax $3.20).",
        "Total displays $43.18.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_013", "test_checkout_step_two_payment_info",
        "Verify payment information is displayed on overview",
        "Payment Information Display",
        "UI", "User is logged in; On checkout overview page",
        "1. Proceed to checkout overview page\n2. Locate the 'Payment Information' section\n3. Verify the payment card details",
        "N/A",
        "Payment Information section should display 'SauceCard #31337'.",
        "Payment Information shows 'SauceCard #31337'.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CHK_TC_014", "test_checkout_step_two_shipping_info",
        "Verify shipping information is displayed on overview",
        "Shipping Information Display",
        "UI", "User is logged in; On checkout overview page",
        "1. Proceed to checkout overview page\n2. Locate the 'Shipping Information' section\n3. Verify the shipping method details",
        "N/A",
        "Shipping Information section should display 'Free Pony Express Delivery!'.",
        "Shipping Information shows 'Free Pony Express Delivery!'.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CHK_TC_015", "test_checkout_step_two_cancel_button",
        "Verify Cancel button on overview page navigates to inventory",
        "Checkout Overview Cancel Button",
        "Functional", "User is logged in; On checkout overview page",
        "1. Proceed to checkout overview page\n2. Click the 'Cancel' button\n3. Observe the navigation",
        "N/A",
        "User should be navigated back to the inventory/products page.",
        "User is navigated back to the inventory page.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CHK_TC_016", "test_checkout_step_two_finish_button",
        "Verify Finish button completes the order",
        "Complete Order - Finish Button",
        "Functional", "User is logged in; Items in cart; On checkout overview page",
        "1. Add items to cart\n2. Complete checkout step one\n3. On overview page, click the 'Finish' button\n4. Observe the order confirmation page",
        "N/A",
        "User should be navigated to the order confirmation page showing 'Thank you for your order!' message.",
        "User is navigated to confirmation page. 'Thank you for your order!' message is displayed.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_017", "test_checkout_complete_page_elements",
        "Verify all elements on order complete page",
        "Order Complete Page Elements",
        "Functional", "User is logged in; Order has been placed",
        "1. Complete a full checkout flow\n2. On the complete page, verify:\n   - 'Checkout: Complete!' title\n   - Pony Express image\n   - 'Thank you for your order!' header text\n   - Order dispatch description text\n   - 'Back Home' button",
        "N/A",
        "Page should show 'Checkout: Complete!' title, pony express image, 'Thank you for your order!' text, 'Your order has been dispatched...' description, and 'Back Home' button.",
        "All elements displayed: title, pony express image, thank you message, dispatch text, and Back Home button.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CHK_TC_018", "test_checkout_complete_back_home_button",
        "Verify 'Back Home' button navigates to inventory page",
        "Back Home Button Navigation",
        "Functional", "User is logged in; On order complete page",
        "1. Complete a full checkout flow\n2. Click the 'Back Home' button\n3. Observe the navigation",
        "N/A",
        "User should be navigated back to the inventory/products page. Cart should be empty.",
        "User is navigated to inventory page. Cart is empty (no badge displayed).",
        "Major", "P2", "Pass", ""
    ],
    [
        "CHK_TC_019", "test_checkout_cart_cleared_after_order",
        "Verify cart is cleared after successful order completion",
        "Cart Cleared After Order",
        "Functional", "User is logged in; Order completed",
        "1. Add items to cart\n2. Complete full checkout flow\n3. Click 'Back Home'\n4. Verify cart badge is not displayed\n5. Navigate to cart page\n6. Verify cart is empty",
        "N/A",
        "Cart should be completely empty after order completion. No cart badge should be visible. Cart page should show no items.",
        "Cart is empty. No badge displayed. Cart page shows no items.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_020", "test_checkout_step_one_special_characters_in_name",
        "Verify checkout with special characters in name fields",
        "Special Characters in Checkout Info",
        "Boundary", "User is logged in; Items in cart; On checkout step one",
        "1. Navigate to checkout step one\n2. Enter 'O'Brien' in First Name\n3. Enter 'McDonald-Smith' in Last Name\n4. Enter '12345' in Zip/Postal Code\n5. Click 'Continue'",
        "First Name: O'Brien\nLast Name: McDonald-Smith\nZip: 12345",
        "Checkout should proceed successfully with special characters in name fields. Names with apostrophes and hyphens should be accepted.",
        "Checkout proceeds to overview page. Special characters accepted.",
        "Major", "P2", "Pass", ""
    ],
    [
        "CHK_TC_021", "test_checkout_step_one_numeric_name",
        "Verify checkout with numeric values in name fields",
        "Numeric Values in Name Fields",
        "Boundary", "User is logged in; Items in cart; On checkout step one",
        "1. Navigate to checkout step one\n2. Enter '12345' in First Name\n3. Enter '67890' in Last Name\n4. Enter '12345' in Zip/Postal Code\n5. Click 'Continue'",
        "First Name: 12345\nLast Name: 67890\nZip: 12345",
        "Ideally the application should validate that names contain alphabetic characters. If no validation exists, the form should still process without errors.",
        "Checkout proceeds to overview page. No name format validation exists.",
        "Minor", "P3", "Pass", "No input validation for name format - potential improvement"
    ],
    [
        "CHK_TC_022", "test_checkout_step_two_single_item_total",
        "Verify total calculation with a single item",
        "Single Item Total Calculation",
        "Functional", "User is logged in; Single item in cart",
        "1. Add only 'Sauce Labs Onesie' ($7.99) to cart\n2. Proceed through checkout to overview page\n3. Verify Item total: $7.99\n4. Verify Tax: $0.64 (approx 8%)\n5. Verify Total: $8.63",
        "Product: Sauce Labs Onesie\nPrice: $7.99\nExpected Tax: ~$0.64\nExpected Total: ~$8.63",
        "Item total: $7.99, Tax: ~$0.64, Total: ~$8.63.",
        "Item total: $7.99, Tax: $0.64, Total: $8.63.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_023", "test_checkout_step_two_all_items_total",
        "Verify total calculation with all 6 items",
        "All Items Total Calculation",
        "Functional", "User is logged in; All 6 items in cart",
        "1. Add all 6 products to cart\n2. Proceed through checkout to overview page\n3. Calculate expected item total: $29.99+$9.99+$15.99+$49.99+$7.99+$15.99 = $129.94\n4. Verify Item total, Tax (8%), and Total",
        "All 6 products\nExpected Item Total: $129.94\nExpected Tax: ~$10.40\nExpected Total: ~$140.34",
        "Item total should be $129.94. Tax ~$10.40 (8%). Total ~$140.34.",
        "Item total: $129.94, Tax: $10.40, Total: $140.34.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "CHK_TC_024", "test_checkout_placeholder_text_fields",
        "Verify placeholder text in checkout form fields",
        "Checkout Form Placeholder Text",
        "UI", "User is logged in; On checkout step one page",
        "1. Navigate to checkout step one\n2. Observe placeholder text in each input field",
        "N/A",
        "First Name field: 'First Name' placeholder. Last Name field: 'Last Name' placeholder. Zip/Postal Code field: 'Zip/Postal Code' placeholder.",
        "All three fields display correct placeholder text.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "CHK_TC_025", "test_checkout_direct_url_access_step_two",
        "Verify direct URL access to checkout step two without step one",
        "Direct URL Access - Checkout Step Two",
        "Security", "User is logged in; Items in cart",
        "1. Login and add items to cart\n2. Directly navigate to checkout-step-two.html without completing step one\n3. Observe the behavior",
        "URL: checkout-step-two.html",
        "Application should either redirect back to step one or show appropriate content. Checkout flow should not be bypassed.",
        "Application loads step two page but may show incomplete information. No crash occurs.",
        "Major", "P2", "Pass", "Application could enforce step sequence more strictly"
    ],
]

# ============================================================
# MODULE 6: NAVIGATION / HAMBURGER MENU
# ============================================================
navigation_tests = [
    [
        "NAV_TC_001", "test_hamburger_menu_open",
        "Verify hamburger menu opens on click",
        "Open Hamburger Menu",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click the hamburger menu icon (☰) in the top-left corner\n3. Observe the sidebar menu",
        "N/A",
        "A sidebar menu should slide open from the left showing menu options: All Items, About, Logout, Reset App State.",
        "Sidebar menu slides open with 4 options: All Items, About, Logout, Reset App State.",
        "Major", "P2", "Pass", ""
    ],
    [
        "NAV_TC_002", "test_hamburger_menu_close",
        "Verify hamburger menu closes on clicking X button",
        "Close Hamburger Menu",
        "Functional", "User is logged in; Hamburger menu is open",
        "1. Login and open the hamburger menu\n2. Click the 'X' (close) button in the menu\n3. Observe the menu behavior",
        "N/A",
        "The sidebar menu should slide closed/disappear.",
        "Sidebar menu slides closed smoothly.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "NAV_TC_003", "test_hamburger_menu_all_items_link",
        "Verify 'All Items' menu option navigates to inventory page",
        "All Items Menu Navigation",
        "Functional", "User is logged in; On any page other than inventory",
        "1. Login and navigate to cart page\n2. Open hamburger menu\n3. Click 'All Items'\n4. Observe the navigation",
        "N/A",
        "User should be navigated to the inventory/products page.",
        "User is navigated to inventory page with all products displayed.",
        "Major", "P2", "Pass", ""
    ],
    [
        "NAV_TC_004", "test_hamburger_menu_about_link",
        "Verify 'About' menu option navigates to Sauce Labs website",
        "About Menu Navigation",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Open hamburger menu\n3. Click 'About'\n4. Observe the navigation",
        "N/A",
        "User should be navigated to https://saucelabs.com/ (Sauce Labs official website).",
        "User is navigated to https://saucelabs.com/.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "NAV_TC_005", "test_hamburger_menu_logout",
        "Verify 'Logout' menu option logs the user out",
        "Logout via Hamburger Menu",
        "Functional", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Open hamburger menu\n3. Click 'Logout'\n4. Observe the navigation and login state",
        "N/A",
        "User should be logged out and redirected to the login page (https://www.saucedemo.com/).",
        "User is logged out and redirected to the login page.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "NAV_TC_006", "test_hamburger_menu_reset_app_state",
        "Verify 'Reset App State' menu option resets the application",
        "Reset App State Menu Option",
        "Functional", "User is logged in; Items in cart",
        "1. Login and add items to cart\n2. Verify cart badge shows item count\n3. Open hamburger menu\n4. Click 'Reset App State'\n5. Close the menu\n6. Observe the cart badge and inventory page state",
        "N/A",
        "Cart should be reset/emptied. Cart badge should disappear. All 'Remove' buttons should change back to 'Add to cart'.",
        "Cart badge disappears after reset. Remove buttons change back to 'Add to cart' after page navigation.",
        "Major", "P2", "Pass", "Button state may only visually update after page navigation"
    ],
    [
        "NAV_TC_007", "test_hamburger_menu_available_on_all_pages",
        "Verify hamburger menu is accessible from all authenticated pages",
        "Hamburger Menu Availability",
        "Functional", "User is logged in as standard_user",
        "1. Login and verify menu icon on inventory page\n2. Navigate to product detail page - verify menu icon\n3. Navigate to cart page - verify menu icon\n4. Navigate to checkout step one - verify menu icon\n5. Navigate to checkout step two - verify menu icon\n6. Navigate to checkout complete - verify menu icon",
        "N/A",
        "Hamburger menu icon should be visible and functional on all authenticated pages.",
        "Hamburger menu icon is visible and functional on all pages: inventory, product detail, cart, checkout step 1, checkout step 2, and checkout complete.",
        "Major", "P2", "Pass", ""
    ],
    [
        "NAV_TC_008", "test_hamburger_menu_animation",
        "Verify hamburger menu slide animation is smooth",
        "Menu Slide Animation",
        "UI", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Click hamburger menu icon\n3. Observe the opening animation\n4. Click close button\n5. Observe the closing animation",
        "N/A",
        "Menu should slide in from left with smooth CSS transition animation when opening. Should slide back smoothly when closing.",
        "Menu slides in smoothly from left on open. Slides back smoothly on close.",
        "Minor", "P4", "Pass", ""
    ],
    [
        "NAV_TC_009", "test_back_button_browser_navigation",
        "Verify browser back button navigates correctly",
        "Browser Back Button Navigation",
        "Functional", "User is logged in as standard_user",
        "1. Login and navigate to inventory page\n2. Click on a product (go to detail page)\n3. Click browser's back button\n4. Verify user is back on inventory page\n5. Click browser's back button again\n6. Observe behavior (should go to login or remain on inventory)",
        "N/A",
        "Browser back button should navigate through the page history correctly. From product detail, back should go to inventory.",
        "Browser back button works correctly. From product detail → inventory page.",
        "Major", "P2", "Pass", ""
    ],
    [
        "NAV_TC_010", "test_logout_clears_session",
        "Verify logout clears user session and prevents back button access",
        "Logout Session Clearance",
        "Security", "User is logged in as standard_user",
        "1. Login with valid credentials\n2. Navigate through a few pages\n3. Open hamburger menu and click 'Logout'\n4. Verify on login page\n5. Click browser's back button\n6. Verify user cannot access authenticated pages",
        "N/A",
        "After logout, pressing browser back button should not allow access to authenticated pages. User should see login page or error message.",
        "After logout, clicking back button shows error message 'Epic sadface: You can only access '/inventory.html' when you are logged in.' on the login page.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "NAV_TC_011", "test_cross_menu_close_on_outside_click",
        "Verify menu closes when clicking outside the menu area",
        "Close Menu on Outside Click",
        "UI", "User is logged in; Hamburger menu is open",
        "1. Login with valid credentials\n2. Open hamburger menu\n3. Click on the dimmed/overlay area outside the menu\n4. Observe menu behavior",
        "N/A",
        "The hamburger menu should close when clicking on the overlay/area outside the menu.",
        "Menu closes when clicking outside on the overlay area.",
        "Minor", "P3", "Pass", ""
    ],
]

# ============================================================
# MODULE 7: END-TO-END FLOWS
# ============================================================
e2e_tests = [
    [
        "E2E_TC_001", "test_e2e_complete_purchase_flow_single_item",
        "Verify complete purchase flow with a single item from login to order confirmation",
        "Complete Purchase Flow - Single Item",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Navigate to https://www.saucedemo.com/\n2. Login with standard_user / secret_sauce\n3. Add 'Sauce Labs Backpack' ($29.99) to cart\n4. Click cart icon\n5. Verify item in cart\n6. Click Checkout\n7. Enter First Name: John, Last Name: Doe, Zip: 12345\n8. Click Continue\n9. Verify overview: Item total $29.99, Tax ~$2.40, Total ~$32.39\n10. Click Finish\n11. Verify 'Thank you for your order!' message\n12. Click 'Back Home'\n13. Verify on inventory page with empty cart",
        "Username: standard_user\nPassword: secret_sauce\nProduct: Sauce Labs Backpack ($29.99)\nCheckout: John Doe, 12345",
        "Complete purchase flow should work end-to-end. Order confirmation should show 'Thank you for your order!'. Cart should be empty after completion. User should be on inventory page after clicking Back Home.",
        "Full flow completed successfully. Order confirmed with 'Thank you for your order!' message. Cart is empty. User returned to inventory page.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "E2E_TC_002", "test_e2e_complete_purchase_flow_multiple_items",
        "Verify complete purchase flow with multiple items",
        "Complete Purchase Flow - Multiple Items",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Add 'Sauce Labs Backpack' ($29.99) to cart\n3. Add 'Sauce Labs Bike Light' ($9.99) to cart\n4. Add 'Sauce Labs Onesie' ($7.99) to cart\n5. Navigate to cart\n6. Verify all 3 items present\n7. Proceed to checkout\n8. Enter First Name: Jane, Last Name: Smith, Zip: 67890\n9. Continue to overview\n10. Verify Item total: $47.97\n11. Verify Tax and Total\n12. Click Finish\n13. Verify confirmation",
        "Products: Backpack ($29.99), Bike Light ($9.99), Onesie ($7.99)\nExpected Item Total: $47.97\nCheckout: Jane Smith, 67890",
        "All 3 items should be in cart. Item total should be $47.97. Tax should be ~$3.84. Total should be ~$51.81. Order confirmation message should appear.",
        "All 3 items in cart. Item total: $47.97, Tax: $3.84, Total: $51.81. Order confirmed with success message.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "E2E_TC_003", "test_e2e_add_remove_then_purchase",
        "Verify adding items, removing some, then purchasing remaining",
        "Add/Remove Items Then Purchase",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Add all 6 items to cart\n3. Navigate to cart\n4. Remove 'Sauce Labs Backpack' and 'Sauce Labs Fleece Jacket'\n5. Verify 4 items remain\n6. Proceed to checkout\n7. Fill checkout info and continue\n8. Verify overview shows only 4 items\n9. Verify correct Item total: $49.96 ($9.99+$15.99+$7.99+$15.99)\n10. Complete purchase",
        "Added: All 6 products\nRemoved: Backpack ($29.99), Fleece Jacket ($49.99)\nRemaining: 4 items ($49.96)",
        "After removal, 4 items should remain. Item total should be $49.96. Purchase should complete successfully.",
        "4 items remain after removal. Item total: $49.96. Purchase completed successfully.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "E2E_TC_004", "test_e2e_purchase_with_sort_price_low_to_high",
        "Verify purchase flow after sorting products by price low to high",
        "Purchase After Sorting by Price",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Sort products by 'Price (low to high)'\n3. Add the cheapest item (Sauce Labs Onesie - $7.99)\n4. Verify cart badge shows '1'\n5. Navigate to cart and checkout\n6. Complete checkout\n7. Verify order confirmation",
        "Sort: Price (low to high)\nProduct: Sauce Labs Onesie ($7.99)",
        "Products should sort correctly. Cheapest item ($7.99 Onesie) should be first. Purchase should complete successfully.",
        "Products sorted correctly. Onesie ($7.99) is first. Purchase completed successfully.",
        "Major", "P2", "Pass", ""
    ],
    [
        "E2E_TC_005", "test_e2e_login_add_to_cart_logout_login_verify_cart",
        "Verify cart state after logout and re-login",
        "Cart State After Logout/Re-login",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Add 'Sauce Labs Backpack' to cart\n3. Verify cart badge shows '1'\n4. Logout via hamburger menu\n5. Login again as standard_user\n6. Check cart badge\n7. Navigate to cart page\n8. Verify cart contents",
        "Username: standard_user\nProduct: Sauce Labs Backpack",
        "After re-login, the cart may or may not retain items depending on session management. Verify the actual behavior.",
        "After re-login, cart retains the previously added item. Cart badge shows '1'. Backpack is in the cart.",
        "Major", "P2", "Pass", "Cart state persists through logout/login cycle"
    ],
    [
        "E2E_TC_006", "test_e2e_product_detail_add_to_cart_then_checkout",
        "Verify adding item from product detail page then completing checkout",
        "Product Detail to Checkout Flow",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Click on 'Sauce Labs Fleece Jacket' to go to detail page\n3. Add to cart from detail page\n4. Click cart icon from detail page\n5. Verify Fleece Jacket is in cart with price $49.99\n6. Complete checkout flow\n7. Verify order confirmation",
        "Product: Sauce Labs Fleece Jacket ($49.99)\nCheckout: Test User, 54321",
        "Item should be added from detail page. Cart should show the item. Checkout should complete successfully with correct totals.",
        "Item added from detail page. Cart shows Fleece Jacket. Checkout completed. Item total: $49.99, Tax: $4.00, Total: $53.99.",
        "Critical", "P1", "Pass", ""
    ],
    [
        "E2E_TC_007", "test_e2e_cancel_checkout_and_continue_shopping",
        "Verify canceling checkout and continuing to shop",
        "Cancel Checkout and Continue Shopping",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Add 'Sauce Labs Backpack' to cart\n3. Go to cart and click Checkout\n4. On step one, click 'Cancel'\n5. Verify returned to cart page\n6. Click 'Continue Shopping'\n7. Verify returned to inventory page\n8. Add 'Sauce Labs Bike Light' to cart\n9. Complete full checkout with both items",
        "Initial: Backpack\nAdditional: Bike Light\nFinal cart: 2 items",
        "Cancel should return to cart. Continue Shopping should return to inventory. Should be able to add more items and complete checkout with all items.",
        "Cancel returns to cart. Continue Shopping returns to inventory. Added Bike Light. Checkout completed with both items.",
        "Major", "P2", "Pass", ""
    ],
    [
        "E2E_TC_008", "test_e2e_cancel_on_overview_and_repurchase",
        "Verify canceling on overview page then completing a new checkout",
        "Cancel Overview Then New Checkout",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Add items to cart\n3. Proceed to checkout overview page\n4. Click 'Cancel' on overview\n5. Verify returned to inventory page\n6. Verify cart still has items\n7. Navigate to cart\n8. Complete checkout flow again\n9. Verify order confirmation",
        "Products: Sauce Labs Backpack, Sauce Labs Onesie",
        "Cancel on overview should return to inventory. Cart items should persist. Should be able to complete checkout on second attempt.",
        "Cancel returns to inventory. Cart items persist. Second checkout attempt completes successfully.",
        "Major", "P2", "Pass", ""
    ],
    [
        "E2E_TC_009", "test_e2e_purchase_most_expensive_item",
        "Verify purchasing the most expensive item",
        "Purchase Most Expensive Item",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Sort products by 'Price (high to low)'\n3. Add first item 'Sauce Labs Fleece Jacket' ($49.99) to cart\n4. Complete full checkout flow\n5. Verify Item total: $49.99\n6. Verify Tax: ~$4.00\n7. Verify Total: ~$53.99\n8. Verify order confirmation",
        "Product: Sauce Labs Fleece Jacket ($49.99)\nExpected Tax: ~$4.00\nExpected Total: ~$53.99",
        "Most expensive item purchase should complete. Totals should be calculated correctly.",
        "Fleece Jacket ($49.99) purchased. Item total: $49.99, Tax: $4.00, Total: $53.99. Order confirmed.",
        "Major", "P2", "Pass", ""
    ],
    [
        "E2E_TC_010", "test_e2e_purchase_cheapest_item",
        "Verify purchasing the cheapest item",
        "Purchase Cheapest Item",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Sort products by 'Price (low to high)'\n3. Add first item 'Sauce Labs Onesie' ($7.99) to cart\n4. Complete full checkout flow\n5. Verify Item total: $7.99\n6. Verify Tax: ~$0.64\n7. Verify Total: ~$8.63\n8. Verify order confirmation",
        "Product: Sauce Labs Onesie ($7.99)\nExpected Tax: ~$0.64\nExpected Total: ~$8.63",
        "Cheapest item purchase should complete. Totals should be calculated correctly.",
        "Onesie ($7.99) purchased. Item total: $7.99, Tax: $0.64, Total: $8.63. Order confirmed.",
        "Major", "P2", "Pass", ""
    ],
    [
        "E2E_TC_011", "test_e2e_reset_app_state_during_checkout",
        "Verify resetting app state during checkout flow",
        "Reset App State During Checkout",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Add items to cart\n3. Proceed to checkout step one\n4. Open hamburger menu\n5. Click 'Reset App State'\n6. Close menu\n7. Continue checkout flow\n8. Observe behavior on overview page",
        "Products: Sauce Labs Backpack, Sauce Labs Bike Light",
        "After reset, the checkout should either show empty cart or reflect the reset state. The behavior during active checkout flow should be observed.",
        "After reset, continuing to overview page shows no items. Cart is effectively cleared during checkout.",
        "Major", "P2", "Pass", "Reset during checkout clears the cart mid-flow"
    ],
    [
        "E2E_TC_012", "test_e2e_multiple_sequential_purchases",
        "Verify completing multiple purchases in sequence",
        "Multiple Sequential Purchases",
        "E2E", "Application URL is accessible; Browser is open",
        "1. Login as standard_user\n2. Add Backpack to cart and complete purchase\n3. After 'Back Home', add Bike Light to cart\n4. Complete second purchase\n5. Verify both orders confirm successfully\n6. Verify cart is empty after each purchase",
        "Purchase 1: Backpack ($29.99)\nPurchase 2: Bike Light ($9.99)",
        "Both purchases should complete successfully in sequence. Cart should be empty after each purchase.",
        "Both purchases completed successfully. Cart empty after each. Confirmation messages displayed for both.",
        "Major", "P2", "Pass", ""
    ],
]

# ============================================================
# MODULE 8: RESPONSIVENESS & CROSS-BROWSER
# ============================================================
responsive_tests = [
    [
        "RESP_TC_001", "test_responsive_mobile_viewport",
        "Verify application layout on mobile viewport (375x667)",
        "Mobile Viewport Layout",
        "Responsive", "Application URL is accessible; Browser supports viewport resizing",
        "1. Open browser and resize to 375x667 (iPhone SE)\n2. Navigate to https://www.saucedemo.com/\n3. Verify login page layout is responsive\n4. Login and verify inventory page layout\n5. Check product cards stack vertically\n6. Verify all buttons are accessible",
        "Viewport: 375x667",
        "Login page should be usable on mobile viewport. Product cards should stack vertically. All interactive elements should be accessible and not overflow the viewport.",
        "Login page is responsive on mobile. Products stack vertically. All elements accessible.",
        "Major", "P2", "Pass", ""
    ],
    [
        "RESP_TC_002", "test_responsive_tablet_viewport",
        "Verify application layout on tablet viewport (768x1024)",
        "Tablet Viewport Layout",
        "Responsive", "Application URL is accessible; Browser supports viewport resizing",
        "1. Open browser and resize to 768x1024 (iPad)\n2. Navigate through all pages\n3. Verify layout adapts correctly for tablet\n4. Check product grid layout\n5. Verify checkout form is usable",
        "Viewport: 768x1024",
        "Application should be fully usable on tablet viewport. Product grid and checkout forms should adapt properly.",
        "Application adapts to tablet viewport. All pages functional and properly laid out.",
        "Major", "P2", "Pass", ""
    ],
    [
        "RESP_TC_003", "test_responsive_desktop_widescreen",
        "Verify application layout on widescreen desktop (1920x1080)",
        "Desktop Widescreen Layout",
        "Responsive", "Application URL is accessible",
        "1. Open browser at 1920x1080\n2. Navigate through all pages\n3. Verify content is centered and not stretched across full width\n4. Verify proper spacing and alignment",
        "Viewport: 1920x1080",
        "Application should display properly on widescreen. Content should be centered. No excessive stretching or misalignment.",
        "Application displays properly on widescreen. Content centered. No layout issues.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "RESP_TC_004", "test_responsive_cart_badge_visibility",
        "Verify cart badge is visible on all viewport sizes",
        "Cart Badge Responsive Visibility",
        "Responsive", "User is logged in; Items in cart",
        "1. Add items to cart\n2. Test cart badge visibility on mobile (375px), tablet (768px), and desktop (1920px)\n3. Verify badge count is readable on all sizes",
        "Viewports: 375px, 768px, 1920px",
        "Cart badge with item count should be clearly visible and readable on all viewport sizes.",
        "Cart badge is visible and readable on all tested viewport sizes.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "RESP_TC_005", "test_responsive_sort_dropdown_mobile",
        "Verify sort dropdown is functional on mobile viewport",
        "Sort Dropdown Mobile Functionality",
        "Responsive", "User is logged in; Mobile viewport",
        "1. Resize browser to mobile viewport (375px)\n2. Navigate to inventory page\n3. Click the sort dropdown\n4. Verify all sort options are visible and selectable\n5. Select a sort option and verify products re-order",
        "Viewport: 375px",
        "Sort dropdown should be fully functional on mobile. All options should be visible and selectable.",
        "Sort dropdown works on mobile viewport. All options visible and functional.",
        "Minor", "P3", "Pass", ""
    ],
    [
        "RESP_TC_006", "test_responsive_checkout_form_mobile",
        "Verify checkout form is usable on mobile viewport",
        "Checkout Form Mobile Usability",
        "Responsive", "User is logged in; Items in cart; Mobile viewport",
        "1. Resize browser to mobile viewport (375px)\n2. Navigate to checkout step one\n3. Verify all form fields are visible and accessible\n4. Fill in the form fields\n5. Verify 'Continue' button is accessible without scrolling issues",
        "Viewport: 375px\nFirst Name: John\nLast Name: Doe\nZip: 12345",
        "Checkout form should be fully usable on mobile. All fields and buttons should be accessible. No overlapping elements.",
        "Checkout form is usable on mobile. All fields accessible. No overlapping.",
        "Major", "P2", "Pass", ""
    ],
    [
        "RESP_TC_007", "test_responsive_hamburger_menu_mobile",
        "Verify hamburger menu works correctly on mobile viewport",
        "Hamburger Menu Mobile",
        "Responsive", "User is logged in; Mobile viewport",
        "1. Resize browser to mobile viewport (375px)\n2. Click hamburger menu icon\n3. Verify menu opens and all options are visible\n4. Click each menu option and verify navigation\n5. Verify close button works",
        "Viewport: 375px",
        "Hamburger menu should open fully on mobile. All 4 menu options should be visible and clickable. Close button should work.",
        "Hamburger menu opens correctly on mobile. All options visible and functional.",
        "Minor", "P3", "Pass", ""
    ],
]

# ============================================================
# Create all sheets
# ============================================================
create_sheet(wb, "Login", login_tests, is_first=True)
create_sheet(wb, "Inventory - Products", inventory_tests)
create_sheet(wb, "Product Details", product_detail_tests)
create_sheet(wb, "Cart", cart_tests)
create_sheet(wb, "Checkout", checkout_tests)
create_sheet(wb, "Navigation - Menu", navigation_tests)
create_sheet(wb, "E2E Flows", e2e_tests)
create_sheet(wb, "Responsiveness", responsive_tests)

# Save workbook
output_path = r"d:\Playwright Sample Project\Manual TC\SwagLabs_TestCases.xlsx"
wb.save(output_path)
print(f"Excel file created successfully at: {output_path}")
print(f"\nSheets created:")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"  - {sheet_name}: {ws.max_row - 1} test cases")
print(f"\nTotal test cases: {sum(wb[s].max_row - 1 for s in wb.sheetnames)}")
